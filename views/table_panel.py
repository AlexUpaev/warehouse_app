# views/table_panel.py
import sys
import os
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Any, Optional

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QTableView, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QTableWidget, QTableWidgetItem,
    QHeaderView, QComboBox, QMessageBox, QDialog, QFormLayout, QLineEdit, QDialogButtonBox,
    QScrollArea, QStyledItemDelegate, QFileDialog, QCheckBox, QGroupBox, QRadioButton,
    QButtonGroup, QProgressBar
)
from PySide6.QtCore import Qt, QSortFilterProxyModel, QPoint, Signal, QPropertyAnimation, QEasingCurve
from PySide6.QtGui import QFont, QStandardItemModel, QStandardItem, QValidator, QColor

from database import Database
from utils.permissions import get_user_permissions
from dateutil import tz

# ============================================================================
# КОНСТАНТЫ И МАППИНГИ
# ============================================================================
TABLES = {
    "Пользователи": "users",
    "Категории": "categories",
    "Поставщики": "suppliers",
    "Материалы": "materials",
    "Транзакции": "transactions",
    "История изменений": "material_history"
}

HEADERS = {
    "users": ["ID", "Логин", "ФИО", "Email", "Роль", "Активен", "Создан", "Последний вход"],
    "categories": ["ID", "Название", "Описание", "Создана"],
    "suppliers": ["ID", "Название", "Контактное лицо", "Телефон", "Email", "Адрес", "Создан"],
    "materials": [
        "ID", "Наименование", "Количество", "Ед.", "Цена (₽)", "Мин. запас",
        "Поставщик", "Описание", "Категория", "Создал", "Создан", "Обновлен"
    ],
    "transactions": [
        "ID", "Материал", "Пользователь", "Количество", "Тип",
        "Документ", "Дата документа", "Примечание", "Создана"
    ],
    "material_history": [
        "ID", "Материал", "Было", "Стало", "Разница",
        "Действие", "Примечание", "Изменил", "Время"
    ]
}

COLUMN_MAPPING = {
    "users": {
        "ID": "id", "Логин": "login", "ФИО": "full_name", "Email": "email",
        "Роль": "role", "Активен": "is_active", "Создан": "created_at", "Последний вход": "last_login"
    },
    "categories": {
        "ID": "id", "Название": "name", "Описание": "description", "Создана": "created_at"
    },
    "suppliers": {
        "ID": "id", "Название": "name", "Контактное лицо": "contact_person",
        "Телефон": "phone", "Email": "email", "Адрес": "address", "Создан": "created_at"
    },
    "materials": {
        "ID": "id", "Наименование": "name", "Количество": "quantity", "Ед.": "unit",
        "Цена (₽)": "price", "Мин. запас": "min_quantity", "Поставщик": "supplier",
        "Описание": "description", "Категория": "category_name", "Создал": "created_by_name",
        "Создан": "created_at", "Обновлен": "updated_at"
    },
    "transactions": {
        "ID": "id", "Материал": "material_name", "Пользователь": "user_name",
        "Количество": "quantity", "Тип": "transaction_type", "Документ": "document_number",
        "Дата документа": "document_date", "Примечание": "notes", "Создана": "created_at"
    },
    "material_history": {
        "ID": "id", "Материал": "material_name", "Было": "old_quantity", "Стало": "new_quantity",
        "Разница": "difference", "Действие": "action_type", "Примечание": "notes",
        "Изменил": "changed_by_name", "Время": "changed_at"
    }
}

HIDDEN_FIELDS = {
    "users": ["ID", "Создан", "Последний вход"],
    "categories": ["ID", "Создана"],
    "suppliers": ["ID", "Создан"],
    "materials": ["ID", "Создал", "Создан", "Обновлен"],
    "transactions": ["ID", "Создана"],
    "material_history": ["ID", "Время"]
}

NUMERIC_COLUMNS = {
    "users": [0], "categories": [0], "suppliers": [0],
    "materials": [0, 2, 4, 5], "transactions": [0, 3], "material_history": [0, 2, 3, 4]
}

DATE_COLUMNS = {
    "users": [6, 7], "categories": [3], "suppliers": [6],
    "materials": [10, 11], "transactions": [6, 8], "material_history": [8]
}

MAX_CELL_LENGTH = 60

# ============================================================================
# ВАЛИДАТОРЫ
# ============================================================================
class PhoneValidator(QValidator):
    def validate(self, text, pos):
        if not text:
            return QValidator.Intermediate, text, pos
        digits = ''.join(filter(str.isdigit, text))
        if len(digits) > 11:
            return QValidator.Invalid, text, pos
        return QValidator.Intermediate, text, pos

class PriceValidator(QValidator):
    def validate(self, text, pos):
        if not text:
            return QValidator.Intermediate, text, pos
        try:
            float(text.replace(',', '.'))
            return QValidator.Acceptable, text, pos
        except ValueError:
            return QValidator.Invalid, text, pos

# ============================================================================
# ДИАЛОГ ВВОДА ДАННЫХ
# ============================================================================
class InputForm(QDialog):
    def __init__(self, table_name, fields_or_values, parent=None):
        super().__init__(parent)
        self.table_name = table_name
        form_layout = QFormLayout()
        self.setLayout(form_layout)

        if isinstance(fields_or_values, list):
            fields = fields_or_values
            values = None
        elif isinstance(fields_or_values, dict):
            fields = list(fields_or_values.keys())
            values = fields_or_values
        else:
            raise TypeError("Неподдерживаемый тип аргумента.")

        hidden = HIDDEN_FIELDS.get(table_name, [])
        fields = [f for f in fields if f not in hidden]

        if table_name == "users" and values is None:
            fields = [f for f in fields if f != "Пароль" and f != "Подтверждение пароля"]
            try:
                email_idx = fields.index("Email")
                fields.insert(email_idx + 1, "Пароль")
                fields.insert(email_idx + 2, "Подтверждение пароля")
            except ValueError:
                fields.extend(["Пароль", "Подтверждение пароля"])

        self.input_fields = {}
        placeholders = {
            "Логин": "Например: user123", "Пароль": "Минимум 6 символов",
            "Подтверждение пароля": "Повторите пароль", "ФИО": "Иванов Иван Иванович",
            "Email": "user@example.com", "Роль": "admin или user",
            "Наименование": "Например: Саморез по дереву 4.2x75", "Количество": "Число, например: 100",
            "Ед.": "шт, кг, м, мешок и т.д.", "Цена (₽)": "Например: 1.50 или 1500",
            "Мин. запас": "Например: 50", "Поставщик": "ООО \"Название\"",
            "Описание": "Краткое описание товара", "Категория": "ID категории или название",
            "Название": "Например: Крепежные изделия", "Телефон": "+7 (999) 123-45-67",
            "Адрес": "г. Москва, ул. Примерная, д. 1", "Контактное лицо": "Иванов Иван",
            "Документ": "ПРИХ-001/2025", "Дата документа": "дд.мм.гггг",
            "Примечание": "Дополнительная информация", "Тип": "incoming, outgoing, adjustment",
            "Пользователь": "ID или ФИО пользователя", "Материал": "ID или название материала",
            "Действие": "add, remove, adjust, create, update", "Было": "Предыдущее количество",
            "Стало": "Новое количество", "Разница": "Разница (Стало - Было)",
            "Время": "дд.мм.гггг чч:мм", "Создан": "дд.мм.гггг чч:мм",
            "Обновлен": "дд.мм.гггг чч:мм", "Последний вход": "дд.мм.гггг чч:мм",
            "Активен": "True или False", "Создал": "ID пользователя", "Изменил": "ID пользователя"
        }

        masks = {"Телефон": "+7 (000) 000-00-00", "Дата документа": "00.00.0000"}

        for field in fields:
            label = QLabel(field)
            input_field = QLineEdit()
            if field in placeholders:
                input_field.setPlaceholderText(placeholders[field])
            if field in masks:
                input_field.setInputMask(masks[field])
            elif field == "Телефон":
                input_field.setValidator(PhoneValidator())
            elif field in ["Цена (₽)", "Было", "Стало", "Разница", "Количество", "Мин. запас"]:
                input_field.setValidator(PriceValidator())
            elif field in ["Пароль", "Подтверждение пароля"]:
                input_field.setEchoMode(QLineEdit.Password)

            if values:
                value = values[field]
                if value is not None:
                    if isinstance(value, (date, datetime)):
                        if isinstance(value, datetime) and value.tzinfo is not None:
                            local_tz = tz.tzlocal()
                            local_value = value.astimezone(local_tz)
                        else:
                            local_value = value
                        input_field.setText(local_value.strftime("%d.%m.%Y %H:%M"))
                    else:
                        input_field.setText(str(value))

            self.input_fields[field] = input_field
            form_layout.addRow(label, input_field)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.validate_and_submit)
        button_box.rejected.connect(self.reject)
        form_layout.addRow(button_box)

    def validate_and_submit(self):
        if self.table_name == "users":
            password = self.input_fields.get("Пароль", QLineEdit()).text()
            confirm = self.input_fields.get("Подтверждение пароля", QLineEdit()).text()
            if len(password) < 6:
                QMessageBox.warning(self, "Ошибка", "Пароль должен быть не менее 6 символов")
                return
            if password != confirm:
                QMessageBox.warning(self, "Ошибка", "Пароли не совпадают")
                return
            if "Подтверждение пароля" in self.input_fields:
                del self.input_fields["Подтверждение пароля"]
        self.accept()

# ============================================================================
# КАСТОМНЫЕ ВИДЖЕТЫ ТАБЛИЦЫ
# ============================================================================
class SortableHeaderView(QHeaderView):
    sortRequested = Signal(int, bool)
    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self.sort_column = -1
        self.sort_ascending = True
        self.setSectionsClickable(True)
        self.sectionClicked.connect(self.on_section_clicked)
        self.setDefaultAlignment(Qt.AlignLeft | Qt.AlignVCenter)

    def on_section_clicked(self, logical_index):
        if logical_index == 0: return
        if self.sort_column == logical_index:
            self.sort_ascending = not self.sort_ascending
        else:
            self.sort_column = logical_index
            self.sort_ascending = True
        self.sortRequested.emit(logical_index, self.sort_ascending)
        self.viewport().update()

class CustomFilterProxyModel(QSortFilterProxyModel):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.search_text = ""

    def set_search_text(self, text):
        self.search_text = text.strip().lower()
        self.invalidateFilter()

    def filterAcceptsRow(self, source_row, source_parent):
        if not self.search_text: return True
        source_model = self.sourceModel()
        if source_model is None: return True
        for column in range(source_model.columnCount()):
            index = source_model.index(source_row, column, source_parent)
            cell_value = str(source_model.data(index, Qt.DisplayRole) or "").lower()
            if self.search_text in cell_value: return True
        return False

class EditableItemDelegate(QStyledItemDelegate):
    def __init__(self, parent=None, table_panel=None):
        super().__init__(parent)
        self.table_panel = table_panel

    def setEditorData(self, editor, index):
        value = index.model().data(index, Qt.EditRole)
        if isinstance(editor, QLineEdit):
            editor.setText(str(value) if value is not None else "")

    def setModelData(self, editor, model, index):
        if isinstance(editor, QLineEdit):
            new_value = editor.text()
            old_value = model.data(index, Qt.DisplayRole)
            if str(new_value) != str(old_value):
                reply = QMessageBox.question(
                    self.table_panel, "Подтверждение изменения",
                    "Желаете изменить?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                )
                if reply == QMessageBox.Yes:
                    model.setData(index, new_value, Qt.EditRole)
                    if self.table_panel:
                        self.table_panel.save_cell_change(index, old_value, new_value)
                else:
                    editor.setText(str(old_value))

# ============================================================================
# БОКОВОЕ МЕНЮ
# ============================================================================
class SideMenu(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_window = parent
        self.setFixedWidth(280)
        self.setup_ui()

    def setup_ui(self):
        # Белый фон меню
        self.setStyleSheet("QWidget { background-color: white; color: black; }")
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # Заголовок меню с синим фоном
        header = QWidget()
        header.setFixedHeight(80)
        header.setStyleSheet("background-color: #1A529C;")
        header_layout = QVBoxLayout(header)
        header_layout.setContentsMargins(20, 20, 20, 20)
        menu_title = QLabel("Меню")
        menu_title.setFont(QFont('Segoe UI', 18, QFont.Weight.Bold))
        menu_title.setStyleSheet("color: white;")
        header_layout.addWidget(menu_title)
        main_layout.addWidget(header)

        self.buttons_container = QWidget()
        self.buttons_container.setStyleSheet("background-color: white;")
        buttons_layout = QVBoxLayout(self.buttons_container)
        buttons_layout.setContentsMargins(0, 0, 0, 0)
        buttons_layout.setSpacing(2)

        menu_items = [
            ("Личный кабинет", self.open_profile),
            ("Приход/расход материала", self.open_materials_flow),
            ("Отчёты по стройкам", self.open_reports),
        ]
        for icon_text, callback in menu_items:
            btn = self.create_menu_button(icon_text, callback)
            buttons_layout.addWidget(btn)
        buttons_layout.addStretch()
        main_layout.addWidget(self.buttons_container)

    def create_menu_button(self, text, callback):
        btn = QPushButton(text)
        btn.setFixedHeight(55)
        # Чёрный текст, белая кнопка, голубая прозрачная выборка
        btn.setStyleSheet("""
        QPushButton { 
            background-color: white; 
            color: black; 
            text-align: left; 
            padding-left: 30px; 
            font-size: 14px; 
            font-weight: 500; 
            border: none; 
            border-left: 4px solid transparent; 
        }
        QPushButton:hover { 
            background-color: rgba(26, 82, 156, 0.15); 
            border-left: 4px solid #1A529C; 
        }
        QPushButton:pressed { 
            background-color: rgba(26, 82, 156, 0.25); 
        }
        """)
        btn.clicked.connect(callback)
        return btn

    def open_profile(self):
        if hasattr(self.parent_window, 'close_menu'): self.parent_window.close_menu()
        try:
            from .profile_page import ProfilePage
            self.profile_window = ProfilePage(self.parent_window.user_data)
            self.profile_window.back_to_table.connect(self.parent_window.show)
            self.profile_window.show()
            self.parent_window.hide()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть профиль:\n{str(e)}")

    def open_materials_flow(self):
        if hasattr(self.parent_window, 'close_menu'): self.parent_window.close_menu()
        try:
            from .material_flow_page import MaterialFlowPage
            self.material_flow_window = MaterialFlowPage(self.parent_window.user_data)
            self.material_flow_window.back_to_table.connect(self.parent_window.show)
            self.material_flow_window.show()
            self.parent_window.hide()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть страницу: {str(e)}")

    def open_reports(self):
        if hasattr(self.parent_window, 'close_menu'): self.parent_window.close_menu()
        try:
            from .reports_page import ReportsPage
            self.reports_window = ReportsPage(self.parent_window.user_data)
            self.reports_window.back_to_table.connect(self.parent_window.show)
            self.reports_window.show()
            self.parent_window.hide()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось открыть отчёты: {str(e)}")

# ============================================================================
# ХЕЛПЕР ЭКСПОРТА/ИМПОРТА
# ============================================================================
class ExportImportHelper:
    @staticmethod
    def convert_value(value):
        if value is None: return ""
        elif isinstance(value, datetime):
            if value.tzinfo is not None: value = value.astimezone(tz.tzlocal())
            return value.strftime("%d.%m.%Y %H:%M")
        elif isinstance(value, date): return value.strftime("%d.%m.%Y")
        elif isinstance(value, bool): return "Да" if value else "Нет"
        return str(value)

    @staticmethod
    def export_to_xlsx(db, tables_to_export: Dict[str, str], filepath: str) -> bool:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            wb = Workbook()
            if 'Sheet' in wb.sheetnames: del wb['Sheet']
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1A529C", end_color="1A529C", fill_type="solid")
            cell_border = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'), top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
            header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

            for display_name, table_name in tables_to_export.items():
                rows = db.get_table_data(table_name)
                if not rows: continue
                ws = wb.create_sheet(title=display_name[:31])
                headers = list(rows[0].keys())
                ws.append(headers)
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font; cell.fill = header_fill; cell.border = cell_border; cell.alignment = header_align
                for row in rows:
                    ws.append([ExportImportHelper.convert_value(row.get(h)) for h in headers])
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=ws.max_row, column=col_idx)
                    cell.border = cell_border; cell.alignment = cell_align
                for col_idx, header in enumerate(headers, 1):
                    max_len = max(len(str(header)), max((len(str(ExportImportHelper.convert_value(r.get(header)))) for r in rows), default=0))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
            wb.save(filepath)
            return True
        except Exception as e:
            print(f"Ошибка экспорта XLSX: {e}"); import traceback; traceback.print_exc()
            return False

    @staticmethod
    def import_from_xlsx(db, table_name: str, filepath: str, skip_duplicates: bool = True) -> Dict[str, Any]:
        stats = {'total': 0, 'imported': 0, 'skipped': 0, 'errors': 0, 'error_messages': []}
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb.active
            excel_headers = [cell.value for cell in ws[1] if cell.value]
            valid_columns = []
            try:
                conn = db.get_connection(); cursor = conn.cursor()
                cursor.execute("SELECT column_name FROM information_schema.columns WHERE table_name = %s AND table_schema = 'public'", (table_name,))
                valid_columns = [row[0] for row in cursor.fetchall()]
                cursor.close(); conn.close()
            except Exception as e:
                stats['error_messages'].append(f"Не удалось прочитать схему БД: {e}"); return stats
            auto_fields = ['id', 'created_at', 'updated_at', 'last_login', 'changed_at']
            importable_columns = [c for c in valid_columns if c not in auto_fields]
            column_map = {}
            for h in excel_headers:
                if h and h in importable_columns: column_map[h] = h
                elif h and h.lower() in importable_columns: column_map[h] = h.lower()
            if not column_map:
                stats['error_messages'].append("Нет совпадающих колонок между файлом и таблицей БД"); return stats
            existing = set()
            if skip_duplicates and table_name in ['materials', 'users', 'categories']:
                unique_field = {'materials': 'name', 'users': 'login', 'categories': 'name'}.get(table_name)
                if unique_field and unique_field in importable_columns:
                    try:
                        conn = db.get_connection(); cursor = conn.cursor()
                        cursor.execute(f"SELECT {unique_field} FROM {table_name}")
                        existing = {str(row[0]).lower() for row in cursor.fetchall()}
                        cursor.close(); conn.close()
                    except: pass
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                stats['total'] += 1
                if all(v is None for v in row): stats['skipped'] += 1; continue
                record = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(excel_headers) and excel_headers[col_idx] in column_map:
                        db_col = column_map[excel_headers[col_idx]]
                        if value is not None and str(value).strip():
                            if db_col in ['quantity', 'min_quantity', 'old_quantity', 'new_quantity', 'difference']:
                                try: value = int(float(str(value).replace(',', '.')))
                                except: pass
                            elif db_col == 'price':
                                try: value = float(str(value).replace(',', '.'))
                                except: pass
                            elif db_col == 'is_active': value = str(value).lower() in ('true', '1', 'да', 'yes')
                            record[db_col] = value
                if skip_duplicates and record:
                    unique_field = {'materials': 'name', 'users': 'login', 'categories': 'name'}.get(table_name)
                    if unique_field and unique_field in record:
                        if str(record[unique_field]).lower() in existing: stats['skipped'] += 1; continue
                if record:
                    try:
                        db.insert_record(table_name, record); stats['imported'] += 1
                        if skip_duplicates and unique_field and unique_field in record: existing.add(str(record[unique_field]).lower())
                    except Exception as e:
                        stats['errors'] += 1; stats['error_messages'].append(f"Строка {row_idx}: {str(e)[:80]}")
            return stats
        except Exception as e:
            stats['errors'] += 1; stats['error_messages'].append(str(e)); import traceback; traceback.print_exc()
            return stats

class ExportImportDialog(QDialog):
    def __init__(self, parent=None, db=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("Экспорт/Импорт данных")
        self.setModal(True)
        self.setMinimumSize(650, 550)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self); layout.setSpacing(15); layout.setContentsMargins(20, 20, 20, 20)
        title = QLabel("Экспорт и импорт данных")
        title.setFont(QFont('Segoe UI', 16, QFont.Weight.Bold)); title.setStyleSheet("color: #1A529C;")
        layout.addWidget(title)
        mode_group = QGroupBox("Режим работы")
        mode_layout = QHBoxLayout(mode_group)
        self.export_radio = QRadioButton("Экспорт данных"); self.import_radio = QRadioButton("Импорт данных")
        self.export_radio.setChecked(True)
        self.mode_group = QButtonGroup(self); self.mode_group.addButton(self.export_radio); self.mode_group.addButton(self.import_radio)
        mode_layout.addWidget(self.export_radio); mode_layout.addWidget(self.import_radio); mode_layout.addStretch()
        layout.addWidget(mode_group)
        self.content_widget = QWidget(); self.content_layout = QVBoxLayout(self.content_widget); self.content_layout.setContentsMargins(0, 0, 0, 0)
        self.setup_export_ui(); self.setup_import_ui()
        self.import_widget.setVisible(False)
        layout.addWidget(self.content_widget)
        self.progress_bar = QProgressBar(); self.progress_bar.setVisible(False); layout.addWidget(self.progress_bar)
        btn_layout = QHBoxLayout(); btn_layout.addStretch()
        self.execute_btn = QPushButton("Экспортировать"); self.execute_btn.setMinimumHeight(40)
        self.execute_btn.setStyleSheet("QPushButton { background-color: #1A529C; color: white; border: none; border-radius: 6px; font-size: 13px; font-weight: 600; padding: 8px 20px; } QPushButton:hover { background-color: #164786; }")
        self.execute_btn.clicked.connect(self.execute); btn_layout.addWidget(self.execute_btn)
        cancel_btn = QPushButton("Отмена"); cancel_btn.setMinimumHeight(40)
        cancel_btn.setStyleSheet("QPushButton { background-color: #6C757D; color: white; border: none; border-radius: 6px; font-size: 13px; font-weight: 600; padding: 8px 20px; } QPushButton:hover { background-color: #5A6268; }")
        cancel_btn.clicked.connect(self.reject); btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
        self.export_radio.toggled.connect(self.on_mode_changed)

    def setup_export_ui(self):
        self.export_widget = QWidget(); exp_layout = QVBoxLayout(self.export_widget); exp_layout.setContentsMargins(0, 0, 0, 0); exp_layout.setSpacing(10)
        tables_group = QGroupBox("Выберите таблицы для экспорта"); tables_layout = QVBoxLayout(tables_group)
        select_all_layout = QHBoxLayout(); self.select_all_cb = QCheckBox("Выбрать все таблицы"); self.select_all_cb.setChecked(True); self.select_all_cb.stateChanged.connect(self.on_select_all)
        select_all_layout.addWidget(self.select_all_cb); select_all_layout.addStretch(); tables_layout.addLayout(select_all_layout)
        self.table_cbs = {}; scroll = QScrollArea(); scroll.setWidgetResizable(True); scroll.setMaximumHeight(200)
        cb_container = QWidget(); cb_layout = QVBoxLayout(cb_container); cb_layout.setContentsMargins(5, 5, 5, 5)
        for name in TABLES.keys():
            cb = QCheckBox(name); cb.setChecked(True); self.table_cbs[name] = cb; cb_layout.addWidget(cb)
        cb_layout.addStretch(); scroll.setWidget(cb_container); tables_layout.addWidget(scroll); exp_layout.addWidget(tables_group)
        fmt_group = QGroupBox("Формат файла"); fmt_layout = QVBoxLayout(fmt_group)
        self.xlsx_radio = QRadioButton("Excel (.xlsx) — рекомендуется"); self.pdf_radio = QRadioButton("PDF (.pdf) — для печати")
        self.xlsx_radio.setChecked(True); fmt_layout.addWidget(self.xlsx_radio); fmt_layout.addWidget(self.pdf_radio); exp_layout.addWidget(fmt_group)
        file_group = QGroupBox("Сохранить как"); file_layout = QHBoxLayout(file_group)
        self.file_label = QLabel("Не выбрано"); self.file_label.setStyleSheet("color: #666;"); file_layout.addWidget(self.file_label, 1)
        browse_btn = QPushButton("Обзор..."); browse_btn.setStyleSheet("QPushButton { background-color: #17A2B8; color: white; border: none; border-radius: 4px; padding: 6px 12px; } QPushButton:hover { background-color: #138496; }")
        browse_btn.clicked.connect(self.browse_export); file_layout.addWidget(browse_btn); exp_layout.addWidget(file_group)
        exp_layout.addStretch(); self.content_layout.addWidget(self.export_widget)

    def setup_import_ui(self):
        self.import_widget = QWidget(); imp_layout = QVBoxLayout(self.import_widget); imp_layout.setContentsMargins(0, 0, 0, 0); imp_layout.setSpacing(10)
        file_group = QGroupBox("Файл для импорта (.xlsx)"); file_layout = QHBoxLayout(file_group)
        self.import_file_label = QLabel("Не выбрано"); self.import_file_label.setStyleSheet("color: #666;"); file_layout.addWidget(self.import_file_label, 1)
        browse_btn = QPushButton("Обзор..."); browse_btn.setStyleSheet("QPushButton { background-color: #17A2B8; color: white; border: none; border-radius: 4px; padding: 6px 12px; } QPushButton:hover { background-color: #138496; }")
        browse_btn.clicked.connect(self.browse_import); file_layout.addWidget(browse_btn); imp_layout.addWidget(file_group)
        table_group = QGroupBox("Целевая таблица"); table_layout = QVBoxLayout(table_group)
        self.import_table_combo = QComboBox(); self.import_table_combo.addItems(TABLES.keys()); table_layout.addWidget(self.import_table_combo); imp_layout.addWidget(table_group)
        opts_group = QGroupBox("Настройки"); opts_layout = QVBoxLayout(opts_group)
        self.skip_dup_cb = QCheckBox("Пропускать дубликаты (по уникальным полям)"); self.skip_dup_cb.setChecked(True); opts_layout.addWidget(self.skip_dup_cb); imp_layout.addWidget(opts_group)
        imp_layout.addStretch(); self.content_layout.addWidget(self.import_widget)

    def on_mode_changed(self):
        if self.export_radio.isChecked():
            self.export_widget.setVisible(True); self.import_widget.setVisible(False); self.execute_btn.setText("Экспортировать")
        else:
            self.export_widget.setVisible(False); self.import_widget.setVisible(True); self.execute_btn.setText("Импортировать")

    def on_select_all(self, state):
        for cb in self.table_cbs.values(): cb.setChecked(state == Qt.Checked)

    def browse_export(self):
        ext = ".xlsx" if self.xlsx_radio.isChecked() else ".pdf"
        flt = "Excel файлы (*.xlsx)" if self.xlsx_radio.isChecked() else "PDF файлы (*.pdf)"
        fp, _ = QFileDialog.getSaveFileName(self, "Сохранить файл", f"export{ext}", flt)
        if fp: self.file_label.setText(fp)

    def browse_import(self):
        fp, _ = QFileDialog.getOpenFileName(self, "Выберите файл", "", "Excel файлы (*.xlsx)")
        if fp: self.import_file_label.setText(fp)

    def execute(self):
        if self.export_radio.isChecked(): self.do_export()
        else: self.do_import()

    def do_export(self):
        selected = {n: t for n, t in TABLES.items() if self.table_cbs[n].isChecked()}
        if not selected: QMessageBox.warning(self, "Ошибка", "Выберите хотя бы одну таблицу"); return
        fp = self.file_label.text()
        if not fp or fp == "Не выбрано": QMessageBox.warning(self, "Ошибка", "Выберите файл для сохранения"); return
        self.progress_bar.setVisible(True); self.progress_bar.setRange(0, 0); self.execute_btn.setEnabled(False); QApplication.processEvents()
        ok = ExportImportHelper.export_to_xlsx(self.db, selected, fp) if self.xlsx_radio.isChecked() else False
        self.progress_bar.setVisible(False); self.execute_btn.setEnabled(True)
        if ok: QMessageBox.information(self, "Успешно", f"Данные экспортированы:\n{fp}"); self.accept()
        else: QMessageBox.critical(self, "Ошибка", "Не удалось экспортировать данные")

    def do_import(self):
        fp = self.import_file_label.text()
        if not fp or fp == "Не выбрано": QMessageBox.warning(self, "Ошибка", "Выберите файл для импорта"); return
        display_name = self.import_table_combo.currentText(); table_name = TABLES[display_name]
        self.progress_bar.setVisible(True); self.progress_bar.setRange(0, 0); self.execute_btn.setEnabled(False); QApplication.processEvents()
        stats = ExportImportHelper.import_from_xlsx(self.db, table_name, fp, self.skip_dup_cb.isChecked())
        self.progress_bar.setVisible(False); self.execute_btn.setEnabled(True)
        msg = f"Статистика импорта в '{display_name}':\n• Всего строк: {stats['total']}\n• Импортировано: {stats['imported']}\n• Пропущено: {stats['skipped']}"
        if stats['errors']: msg += f"\n• Ошибки: {stats['errors']}\n" + "\n".join(stats['error_messages'][:3])
        QMessageBox.information(self, "Результат", msg)
        if stats['imported'] > 0: self.parent().reload_current_table()

# ============================================================================
# ОСНОВНОЙ КЛАСС TablePanel
# ============================================================================
class TablePanel(QMainWindow):
    def __init__(self, user_data):
        super().__init__()
        self.user_data = user_data
        self.db = Database()
        self.current_table_name = None
        self.menu_opened = False
        self.pending_changes = {}
        
        # ЗАГРУЗКА ПРАВ ДОСТУПА
        self.user_role = user_data.get('role', 'user')
        self.user_perms = get_user_permissions(self.user_role)
        
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle(f"Управление складом | {self.user_data['full_name']}")
        self.setGeometry(100, 100, 1100, 750)
        self.setMinimumSize(1024, 768)
        central_widget = QWidget()
        central_widget.setStyleSheet("background-color: white;")
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(0, 0, 0, 0)
        self.content_layout.setSpacing(0)

        # ─── ВЕРХНЯЯ ПАНЕЛЬ ──────────────────────────────────────
        top_panel = QWidget()
        top_panel.setFixedHeight(60)
        top_panel.setStyleSheet("""
            background-color: #F0F8FF; 
            border-bottom: 2px solid #1A529C;
        """)
        top_layout = QHBoxLayout(top_panel)
        top_layout.setContentsMargins(20, 0, 20, 0)
        top_layout.setSpacing(10)

        # Кнопка меню
        self.menu_button = QPushButton("☰")
        self.menu_button.setFixedSize(40, 40)
        self.menu_button.setStyleSheet("""
            QPushButton { 
                background-color: transparent; 
                color: #1A529C; 
                border: none; 
                border-radius: 4px; 
                font-size: 24px; 
                font-weight: bold; 
            } 
            QPushButton:hover { 
                background-color: #E3F2FD; 
                color: #0d47a1; 
            }
        """)
        self.menu_button.clicked.connect(self.toggle_menu)
        top_layout.addWidget(self.menu_button)

        # Информация о пользователе
        user_label = QLabel(f"Пользователь: {self.user_data['full_name']} ({self.user_role})")
        user_label.setFont(QFont('Segoe UI', 14, QFont.Weight.Bold))
        user_label.setStyleSheet("color: #004085;")
        top_layout.addWidget(user_label)

        top_layout.addSpacing(20)

        # Контейнер для кнопок
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(8)

        # Выпадающий список таблиц
        self.table_selector = QComboBox()
        allowed_tables = self.user_perms["tables"]
        self.table_selector.addItems(allowed_tables)
        self.table_selector.setFixedWidth(180)
        self.table_selector.setStyleSheet("""
            QComboBox {
                background-color: white;
                border: 1px solid #1A529C;
                border-radius: 4px;
                padding: 5px 10px;
                font-size: 13px;
            }
            QComboBox:hover {
                border-color: #0d47a1;
            }
            QComboBox::drop-down {
                border: none;
                width: 25px;
            }
        """)
        self.table_selector.currentTextChanged.connect(self.load_table)
        buttons_layout.addWidget(self.table_selector)

        # Кнопка Добавить
        self.add_button = QPushButton("Добавить")
        self.add_button.setFixedSize(100, 32)
        self.add_button.setStyleSheet("""
            QPushButton { 
                background-color: #1A529C; 
                color: white; 
                border: none; 
                border-radius: 4px; 
                font-weight: 600; 
                font-size: 13px; 
            } 
            QPushButton:hover { 
                background-color: #0d47a1; 
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        self.add_button.clicked.connect(self.open_add_form)
        if not self.user_perms["can_edit"]:
            self.add_button.setEnabled(False)
        buttons_layout.addWidget(self.add_button)

        # Кнопка Удалить
        self.delete_button = QPushButton("Удалить")
        self.delete_button.setFixedSize(90, 32)
        self.delete_button.setStyleSheet("""
            QPushButton { 
                background-color: #D32F2F; 
                color: white; 
                border: none; 
                border-radius: 4px; 
                font-weight: 600; 
                font-size: 13px; 
            } 
            QPushButton:hover { 
                background-color: #B71C1C; 
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        self.delete_button.clicked.connect(self.delete_selected_row)
        if not self.user_perms["can_delete"]:
            self.delete_button.setEnabled(False)
        buttons_layout.addWidget(self.delete_button)

        # Поле поиска
        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("Поиск...")
        self.search_bar.setFixedWidth(180)
        self.search_bar.setStyleSheet("""
            QLineEdit {
                background-color: white;
                border: 1px solid #BDC3C7;
                border-radius: 4px;
                padding: 5px 10px;
                font-size: 13px;
            }
            QLineEdit:focus {
                border-color: #1A529C;
            }
        """)
        buttons_layout.addWidget(self.search_bar)

        # Кнопка Поиск
        self.search_button = QPushButton("Поиск")
        self.search_button.setFixedSize(80, 32)
        self.search_button.setStyleSheet("""
            QPushButton { 
                background-color: #17A2B8; 
                color: white; 
                border: none; 
                border-radius: 4px; 
                font-weight: 600; 
                font-size: 13px; 
            } 
            QPushButton:hover { 
                background-color: #138496; 
            }
        """)
        self.search_button.clicked.connect(self.perform_search)
        self.search_bar.returnPressed.connect(self.perform_search)
        buttons_layout.addWidget(self.search_button)

        buttons_layout.addSpacing(10)

        # Кнопка Экспорт
        self.export_btn = QPushButton("Экспорт")
        self.export_btn.setFixedSize(90, 32)
        self.export_btn.setStyleSheet("""
            QPushButton { 
                background-color: #28A745; 
                color: white; 
                border: none; 
                border-radius: 4px; 
                font-weight: 600; 
                font-size: 13px; 
            } 
            QPushButton:hover { 
                background-color: #218838; 
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        self.export_btn.clicked.connect(self.open_export_dialog)
        if not self.user_perms["can_import_export"]:
            self.export_btn.setEnabled(False)
        buttons_layout.addWidget(self.export_btn)

        # Кнопка Импорт
        self.import_btn = QPushButton("Импорт")
        self.import_btn.setFixedSize(90, 32)
        self.import_btn.setStyleSheet("""
            QPushButton { 
                background-color: #FFC107; 
                color: #212529; 
                border: none; 
                border-radius: 4px; 
                font-weight: 600; 
                font-size: 13px; 
            } 
            QPushButton:hover { 
                background-color: #E0A800; 
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        self.import_btn.clicked.connect(self.open_import_dialog)
        if not self.user_perms["can_import_export"]:
            self.import_btn.setEnabled(False)
        buttons_layout.addWidget(self.import_btn)

        # Кнопка Сбросить пароль (только для админа)
        if self.user_perms.get("can_reset_password", False):
            self.reset_password_btn = QPushButton("Сбросить пароль")
            self.reset_password_btn.setFixedSize(140, 32)
            self.reset_password_btn.setStyleSheet("""
                QPushButton { 
                    background-color: #FF9800; 
                    color: white; 
                    border: none; 
                    border-radius: 4px; 
                    font-weight: 600; 
                    font-size: 13px; 
                }
                QPushButton:hover { 
                    background-color: #F57C00; 
                }
                QPushButton:disabled { 
                    background-color: #CCCCCC; 
                    color: #666666; 
                }
            """)
            self.reset_password_btn.setEnabled(False)
            self.reset_password_btn.clicked.connect(self.reset_password_selected_user)
            buttons_layout.addWidget(self.reset_password_btn)

        buttons_layout.addStretch()

        # Кнопка Выйти (крайняя справа)
        logout_button = QPushButton("Выйти")
        logout_button.setFixedSize(80, 32)
        logout_button.setStyleSheet("""
            QPushButton { 
                background-color: #6C757D; 
                color: white; 
                border: none; 
                border-radius: 4px; 
                font-weight: 600; 
                font-size: 13px; 
            } 
            QPushButton:hover { 
                background-color: #5A6268; 
            }
        """)
        logout_button.clicked.connect(self.logout)
        buttons_layout.addWidget(logout_button)

        top_layout.addLayout(buttons_layout)
        self.content_layout.addWidget(top_panel)

        # ─── ТАБЛИЦА ─────────────────────────────────────────────
        self.data_table = QTableView()
        self.data_table.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.data_table.setAlternatingRowColors(True)
        self.data_table.verticalHeader().setVisible(False)
        self.data_table.setSortingEnabled(False)
        
        # Разрешаем редактирование только если есть права
        if self.user_perms["can_edit"]:
            self.data_table.setEditTriggers(QTableView.EditTrigger.DoubleClicked | QTableView.EditTrigger.EditKeyPressed)
        else:
            self.data_table.setEditTriggers(QTableView.EditTrigger.NoEditTriggers)
            
        header = SortableHeaderView(Qt.Horizontal, self.data_table)
        self.data_table.setHorizontalHeader(header)
        header.sortRequested.connect(self.sort_by_column)
        
        # Улучшенный стиль таблицы
        self.data_table.setStyleSheet("""
            QTableView { 
                background-color: white; 
                color: #333333; 
                font-size: 13px; 
                alternate-background-color: #F8F9FA; 
                gridline-color: #E9ECEF;
                border: none;
            }
            QTableView::item {
                padding: 8px;
                border: none;
            }
            QTableView::item:selected {
                background-color: #E3F2FD;
                color: #1A529C;
            }
            QHeaderView::section { 
                background-color: #1A529C; 
                color: white; 
                font-weight: 600; 
                font-size: 13px; 
                padding: 10px 8px; 
                border: none;
                border-bottom: 2px solid #0d47a1;
            }
            QHeaderView::section:hover { 
                background-color: #0d47a1; 
            }
            QHeaderView::section:pressed { 
                background-color: #093a7a; 
            }
        """)
        
        delegate = EditableItemDelegate(self.data_table, self)
        self.data_table.setItemDelegate(delegate)
        self.content_layout.addWidget(self.data_table, 1)

        self.proxy_model = CustomFilterProxyModel()
        self.proxy_model.setSourceModel(None)
        self.proxy_model.setFilterKeyColumn(-1)
        self.proxy_model.setDynamicSortFilter(False)
        self.proxy_model.setSortCaseSensitivity(Qt.CaseInsensitive)

        # Нижняя панель
        bottom_panel = QWidget()
        bottom_panel.setFixedHeight(40)
        bottom_panel.setStyleSheet("""
            background-color: #F5F5F5; 
            border-top: 1px solid #E0E0E0;
        """)
        bottom_layout = QHBoxLayout(bottom_panel)
        bottom_layout.setContentsMargins(20, 0, 20, 0)
        version_label = QLabel("Склад материалов v1.0 | PUTEVI")
        version_label.setStyleSheet("color: #616161; font-size: 12px;")
        bottom_layout.addWidget(version_label)
        bottom_layout.addStretch()
        self.content_layout.addWidget(bottom_panel)
        main_layout.addWidget(self.content_widget)

        # Боковое меню
        self.side_menu = SideMenu(self)
        self.side_menu.setParent(self)
        self.side_menu.setGeometry(-280, 0, 280, self.height())
        self.side_menu.show()
        
        # Оверлей с белой прозрачностью
        self.overlay = QWidget(self)
        self.overlay.setStyleSheet("background-color: rgba(255, 255, 255, 0.3);")
        self.overlay.setGeometry(0, 0, self.width(), self.height())
        self.overlay.hide()
        self.overlay.mousePressEvent = lambda e: self.toggle_menu()
        self.side_menu.raise_()
        
        # Загружаем первую таблицу
        self.load_table(self.table_selector.currentText())
        
        # Подключаем сигнал выбора ПОСЛЕ установки модели
        if self.data_table.selectionModel():
            self.data_table.selectionModel().selectionChanged.connect(self.on_selection_changed)

    # ============================================================================
    # МЕТОДЫ УПРАВЛЕНИЯ ДОСТУПОМ И СБРОСА ПАРОЛЯ
    # ============================================================================
    def on_selection_changed(self):
        """Активирует кнопку сброса пароля только при выборе строки в таблице Пользователи"""
        if not hasattr(self, 'reset_password_btn') or not self.reset_password_btn:
            return
        selected_indexes = self.data_table.selectedIndexes()
        current_table = TABLES[self.table_selector.currentText()]
        if selected_indexes and current_table == "users" and self.user_perms.get("can_reset_password"):
            self.reset_password_btn.setEnabled(True)
        else:
            self.reset_password_btn.setEnabled(False)

    def reset_password_selected_user(self):
        """Открывает диалог сброса пароля"""
        selected_indexes = self.data_table.selectedIndexes()
        if not selected_indexes:
            QMessageBox.warning(self, "Ошибка", "Выберите пользователя для сброса пароля")
            return
        row = selected_indexes[0].row()
        model = self.data_table.model()
        user_id = model.index(row, 0).data()
        current_table = TABLES[self.table_selector.currentText()]
        if current_table != "users":
            QMessageBox.warning(self, "Ошибка", "Сброс пароля доступен только в таблице 'Пользователи'")
            return
        if user_id == self.user_data['id']:
            QMessageBox.information(self, "Смена пароля", "Для изменения собственного пароля используйте раздел 'Личный кабинет'")
            return

        dialog = QDialog(self)
        dialog.setWindowTitle("Сброс пароля пользователя")
        dialog.setModal(True)
        dialog.setMinimumWidth(400)
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        title = QLabel("Сброс пароля")
        title.setFont(QFont('Segoe UI', 14, QFont.Weight.Bold))
        title.setStyleSheet("color: #1A529C;")
        layout.addWidget(title)

        info = QLabel(f"Пользователь ID: {user_id}")
        info.setStyleSheet("color: #666;")
        layout.addWidget(info)

        password_input = QLineEdit()
        password_input.setPlaceholderText("Введите новый пароль")
        password_input.setEchoMode(QLineEdit.Password)
        password_input.setMinimumHeight(40)
        password_input.setStyleSheet("""
            QLineEdit { 
                background-color: #FAFAFA; 
                border: 1px solid #D0D0D0; 
                border-radius: 6px; 
                padding: 10px; 
                font-size: 14px; 
            } 
            QLineEdit:focus { 
                border: 2px solid #1A529C; 
                outline: none; 
            }
        """)
        layout.addWidget(QLabel("Новый пароль:"))
        layout.addWidget(password_input)

        confirm_input = QLineEdit()
        confirm_input.setPlaceholderText("Подтвердите пароль")
        confirm_input.setEchoMode(QLineEdit.Password)
        confirm_input.setMinimumHeight(40)
        confirm_input.setStyleSheet("""
            QLineEdit { 
                background-color: #FAFAFA; 
                border: 1px solid #D0D0D0; 
                border-radius: 6px; 
                padding: 10px; 
                font-size: 14px; 
            } 
            QLineEdit:focus { 
                border: 2px solid #1A529C; 
                outline: none; 
            }
        """)
        layout.addWidget(QLabel("Подтверждение пароля:"))
        layout.addWidget(confirm_input)

        button_layout = QHBoxLayout()
        button_layout.addStretch()
        cancel_btn = QPushButton("Отмена")
        cancel_btn.setFixedHeight(35)
        cancel_btn.setStyleSheet("""
            QPushButton { 
                background-color: #6C757D; 
                color: white; 
                border: none; 
                border-radius: 4px; 
                font-weight: bold; 
                padding: 8px 16px; 
            } 
            QPushButton:hover { 
                background-color: #5A6268; 
            }
        """)
        cancel_btn.clicked.connect(dialog.reject)
        button_layout.addWidget(cancel_btn)

        reset_btn = QPushButton("Сбросить пароль")
        reset_btn.setFixedHeight(35)
        reset_btn.setStyleSheet("""
            QPushButton { 
                background-color: #FF9800; 
                color: white; 
                border: none; 
                border-radius: 4px; 
                font-weight: bold; 
                padding: 8px 16px; 
            } 
            QPushButton:hover { 
                background-color: #F57C00; 
            }
        """)
        reset_btn.clicked.connect(lambda: self._confirm_password_reset(user_id, password_input.text(), confirm_input.text(), dialog))
        button_layout.addWidget(reset_btn)
        layout.addLayout(button_layout)
        dialog.exec()

    def _confirm_password_reset(self, user_id: int, password: str, confirm: str, dialog):
        """Валидация и выполнение сброса пароля"""
        if len(password) < 6:
            QMessageBox.warning(dialog, "Ошибка", "Пароль должен содержать не менее 6 символов")
            return
        if password != confirm:
            QMessageBox.warning(dialog, "Ошибка", "Пароли не совпадают")
            return
        result = self.db.reset_user_password(user_id, password, self.user_data['id'])
        if result['success']:
            QMessageBox.information(dialog, "Успешно", result['message'])
            dialog.accept()
            self.reload_current_table()
        else:
            QMessageBox.critical(dialog, "Ошибка", result['message'])

    # ============================================================================
    # ОСНОВНЫЕ МЕТОДЫ ТАБЛИЦЫ
    # ============================================================================
    def open_export_dialog(self):
        dialog = ExportImportDialog(self, self.db)
        dialog.exec()

    def open_import_dialog(self):
        dialog = ExportImportDialog(self, self.db)
        dialog.import_radio.setChecked(True)
        dialog.on_mode_changed()
        dialog.exec()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if hasattr(self, 'side_menu'): 
            self.side_menu.setGeometry(-280, 0, 280, self.height())
        if hasattr(self, 'overlay'): 
            self.overlay.setGeometry(0, 0, self.width(), self.height())
        self._adjust_column_widths()

    def toggle_menu(self):
        if self.menu_opened: 
            self.close_menu()
        else: 
            self.open_menu()

    def open_menu(self):
        self.overlay.show()
        self.overlay.raise_()
        self.side_menu.raise_()
        self.menu_opened = True
        self.menu_animation = QPropertyAnimation(self.side_menu, b"pos")
        self.menu_animation.setDuration(300)
        self.menu_animation.setEasingCurve(QEasingCurve.InOutCubic)
        self.menu_animation.setStartValue(QPoint(-280, 0))
        self.menu_animation.setEndValue(QPoint(0, 0))
        self.menu_animation.start()

    def close_menu(self):
        self.menu_animation = QPropertyAnimation(self.side_menu, b"pos")
        self.menu_animation.setDuration(300)
        self.menu_animation.setEasingCurve(QEasingCurve.InOutCubic)
        self.menu_animation.setStartValue(QPoint(0, 0))
        self.menu_animation.setEndValue(QPoint(-280, 0))
        self.menu_animation.finished.connect(self.on_menu_closed)
        self.menu_animation.start()

    def on_menu_closed(self):
        self.overlay.hide()
        self.menu_opened = False

    def perform_search(self):
        text = self.search_bar.text()
        self.proxy_model.set_search_text(text)

    def logout(self):
        self.close()
        from .main_window import MainWindow
        self.main_window = MainWindow()
        self.main_window.show()

    def sort_by_column(self, logical_index, ascending):
        table_name = TABLES[self.table_selector.currentText()]
        is_numeric = logical_index in NUMERIC_COLUMNS.get(table_name, [])
        is_date = logical_index in DATE_COLUMNS.get(table_name, [])
        sort_order = Qt.AscendingOrder if ascending else Qt.DescendingOrder
        if is_numeric: 
            self.proxy_model.setSortRole(Qt.EditRole)
        elif is_date: 
            self.proxy_model.setSortRole(Qt.UserRole)
        else: 
            self.proxy_model.setSortCaseSensitivity(Qt.CaseInsensitive)
            self.proxy_model.setSortRole(Qt.DisplayRole)
        self.proxy_model.sort(logical_index, sort_order)
        if isinstance(self.data_table.horizontalHeader(), SortableHeaderView):
            header = self.data_table.horizontalHeader()
            header.sort_column = logical_index
            header.sort_ascending = ascending

    def reset_sort(self):
        self.proxy_model.sort(-1, Qt.AscendingOrder)

    def _adjust_column_widths(self):
        header = self.data_table.horizontalHeader()
        if not header or header.count() == 0: 
            return
        for i in range(header.count()):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.ResizeToContents)
        total_width = sum(header.sectionSize(i) for i in range(header.count()))
        viewport_width = self.data_table.viewport().width()
        if total_width < viewport_width:
            header.setSectionResizeMode(header.count() - 1, QHeaderView.ResizeMode.Stretch)

    def open_add_form(self):
        selected_table = TABLES[self.table_selector.currentText()]
        dialog = InputForm(selected_table, HEADERS[selected_table])
        if dialog.exec() == QDialog.Accepted:
            rus_data = {field: dialog.input_fields[field].text() for field in dialog.input_fields}
            eng_data = {}
            mapping = COLUMN_MAPPING.get(selected_table, {})
            for rus_key, value in rus_data.items():
                if rus_key in HIDDEN_FIELDS.get(selected_table, []): 
                    continue
                eng_key = mapping.get(rus_key, rus_key.lower().replace(" ", "_"))
                if rus_key == "Пароль":
                    from utils.password_helper import PasswordHelper
                    eng_data['password_hash'] = PasswordHelper.hash_password(value)
                elif rus_key == "Активен":
                    eng_data[eng_key] = value.strip().lower() in ('true', '1', 'да', 'yes')
                elif not value or value.strip() == '':
                    eng_data[eng_key] = None
                else:
                    if rus_key in ["Количество", "Мин. запас"]:
                        try: 
                            eng_data[eng_key] = int(value)
                        except: 
                            eng_data[eng_key] = value
                    elif rus_key == "Цена (₽)":
                        try: 
                            eng_data[eng_key] = float(value.replace(',', '.'))
                        except: 
                            eng_data[eng_key] = value
                    else:
                        eng_data[eng_key] = value
            try:
                self.db.insert_record(selected_table, eng_data)
                QMessageBox.information(self, "Успешно", "Запись добавлена!")
                self.reload_current_table()
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Не удалось добавить запись:\n{str(e)}")

    def save_cell_change(self, index, old_value, new_value):
        row = index.row()
        col = index.column()
        table_name = TABLES[self.table_selector.currentText()]
        model = self.data_table.model()
        record_id = model.index(row, 0).data()
        header_name = HEADERS[table_name][col]
        db_column = COLUMN_MAPPING[table_name].get(header_name, header_name.lower().replace(" ", "_"))
        if header_name in ["Количество", "Мин. запас", "ID"]:
            try: 
                new_value = int(new_value)
            except: 
                pass
        elif header_name == "Цена (₽)":
            try: 
                new_value = float(str(new_value).replace(',', '.'))
            except: 
                pass
        elif header_name == "Активен":
            new_value = str(new_value).lower() in ('true', '1', 'да', 'yes')
        try:
            updated_data = {db_column: new_value}
            self.db.update_record(table_name, record_id, updated_data)
            QMessageBox.information(self, "Успешно", "Изменения сохранены в базе данных!")
            self.reload_current_table()
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить изменения:\n{str(e)}")
            model.setData(index, old_value, Qt.DisplayRole)

    def delete_selected_row(self):
        selected_indexes = self.data_table.selectedIndexes()
        if selected_indexes:
            row = selected_indexes[0].row()
            model = self.data_table.model()
            record_id = model.index(row, 0).data()
            table_name = TABLES[self.table_selector.currentText()]
            try:
                dependencies = self.db.get_foreign_key_dependencies(table_name, record_id)
                if dependencies:
                    dep_text = "\n".join([f"• {table}: {count} записей" for table, count in dependencies.items()])
                    total = sum(dependencies.values())
                    msg = QMessageBox(self)
                    msg.setIcon(QMessageBox.Question)
                    msg.setWindowTitle("Подтверждение каскадного удаления")
                    msg.setText(f"Будет удалено {total} связанных записей:")
                    msg.setInformativeText(f"{dep_text}\nЖелаете продолжить?")
                    msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
                    msg.setDefaultButton(QMessageBox.No)
                    result = msg.exec()
                    if result == QMessageBox.Yes:
                        deleted = self.db.cascade_delete(table_name, record_id)
                        deleted_text = "\n".join([f"• {table}: {count}" for table, count in deleted.items()])
                        total_deleted = sum(deleted.values())
                        QMessageBox.information(self, "Успешно", f"Удалено записей: {total_deleted}\n{deleted_text}")
                        self.reload_current_table()
                else:
                    confirm_result = QMessageBox.question(self, "Подтверждение удаления", 
                        "Вы действительно хотите удалить выбранную запись?", 
                        QMessageBox.Yes | QMessageBox.No)
                    if confirm_result == QMessageBox.Yes:
                        self.db.delete_record(table_name, record_id)
                        self.reload_current_table()
                        QMessageBox.information(self, "Успешно", "Запись успешно удалена!")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при удалении:\n{str(e)}")
        else:
            QMessageBox.information(self, "Внимание", "Выберите запись для удаления.")

    def reload_current_table(self):
        self.load_table(self.table_selector.currentText())

    def load_table(self, display_name: str):
        table_name = TABLES[display_name]
        self.current_table_name = table_name
        try:
            rows = self.db.get_table_data(table_name)
            headers = HEADERS[table_name]
            model = QStandardItemModel(len(rows), len(headers))
            model.setHorizontalHeaderLabels(headers)
            for row_idx, row in enumerate(rows):
                row_values = list(row.values())
                for col_idx in range(len(headers)):
                    value = row_values[col_idx] if col_idx < len(row_values) else None
                    item = QStandardItem()
                    if value is None: 
                        display_text = ""
                    elif isinstance(value, datetime):
                        if value.tzinfo is not None:
                            local_value = value.astimezone(tz.tzlocal())
                        else:
                            local_value = value
                        display_text = local_value.strftime("%d.%m.%Y %H:%M")
                    elif isinstance(value, date): 
                        display_text = value.strftime("%d.%m.%Y")
                    elif isinstance(value, bool): 
                        display_text = "true" if value else "false"
                    else:
                        text = str(value)
                        if len(text) > MAX_CELL_LENGTH:
                            text = text[:MAX_CELL_LENGTH - 1] + "…"
                        display_text = text
                    item.setText(display_text)
                    item.setData(display_text, Qt.ItemDataRole.DisplayRole)
                    item.setData(display_text, Qt.ItemDataRole.EditRole)
                    if isinstance(value, (int, float)): 
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
                    else:
                        item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
                    model.setItem(row_idx, col_idx, item)
            self._adjust_column_widths()
            self.proxy_model.setSourceModel(model)
            self.data_table.setModel(self.proxy_model)
            if isinstance(self.data_table.horizontalHeader(), SortableHeaderView):
                header = self.data_table.horizontalHeader()
                header.sort_column = -1
                header.sort_ascending = True
            self.search_bar.clear()
            self.proxy_model.set_search_text("")
            self.reset_sort()
        except Exception as e:
            print(f"Ошибка при загрузке таблицы '{table_name}': {e}")
            import traceback
            traceback.print_exc()
            error_model = QStandardItemModel(1, 1)
            error_model.setHorizontalHeaderLabels(["Ошибка"])
            error_model.setItem(0, 0, QStandardItem(str(e)))
            self.proxy_model.setSourceModel(error_model)
            self.data_table.setModel(self.proxy_model)