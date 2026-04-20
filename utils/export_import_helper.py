import io
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import psycopg2
from psycopg2.extras import RealDictCursor


class ExportImportHelper:
    """Класс для экспорта и импорта данных"""
    
    # Регистрируем шрифт с поддержкой кириллицы
    @staticmethod
    def register_fonts():
        """Регистрирует шрифты для PDF с поддержкой русских символов"""
        try:
            # Пробуем зарегистрировать стандартные шрифты Windows
            font_paths = [
                (r"C:\Windows\Fonts\arial.ttf", "Arial"),
                (r"C:\Windows\Fonts\times.ttf", "Times"),
            ]
            
            for font_path, font_name in font_paths:
                if Path(font_path).exists():
                    pdfmetrics.registerFont(TTFont(font_name, font_path))
                    return font_name
        except:
            pass
        return "Helvetica"  # Fallback
    
    @staticmethod
    def convert_value_for_excel(value):
        """Конвертирует значение для Excel"""
        if value is None:
            return ""
        elif isinstance(value, (datetime, date)):
            return value.strftime("%d.%m.%Y %H:%M") if isinstance(value, datetime) else value.strftime("%d.%m.%Y")
        elif isinstance(value, bool):
            return "Да" if value else "Нет"
        else:
            return str(value)
    
    @staticmethod
    def export_to_xlsx(db, tables_to_export: Dict[str, str], filepath: str) -> bool:
        """
        Экспорт данных в Excel
        
        Args:
            db: объект Database
            tables_to_export: dict {display_name: table_name}
            filepath: путь к файлу
        """
        try:
            wb = Workbook()
            
            # Удаляем стандартный лист
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            # Стили
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1A529C", end_color="1A529C", fill_type="solid")
            cell_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            for display_name, table_name in tables_to_export.items():
                # Получаем данные
                rows = db.get_table_data(table_name)
                if not rows:
                    continue
                
                # Создаем лист
                ws = wb.create_sheet(title=display_name[:31])  # Excel ограничивает имя 31 символом
                
                # Заголовки
                headers = list(rows[0].keys())
                ws.append(headers)
                
                # Стилизация заголовков
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = cell_border
                    cell.alignment = header_alignment
                
                # Данные
                for row_idx, row in enumerate(rows, 2):
                    row_values = [ExportImportHelper.convert_value_for_excel(row.get(h)) for h in headers]
                    ws.append(row_values)
                    
                    # Стилизация ячеек
                    for col_idx in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = cell_border
                        cell.alignment = cell_alignment
                
                # Автоширина колонок
                for col_idx, header in enumerate(headers, 1):
                    max_length = max(
                        len(str(header)),
                        max((len(str(ExportImportHelper.convert_value_for_excel(row.get(header)))) 
                             for row in rows), default=0)
                    )
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
            
            # Сохраняем
            wb.save(filepath)
            return True
            
        except Exception as e:
            print(f"❌ Ошибка экспорта в Excel: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    @staticmethod
    def export_to_pdf(db, tables_to_export: Dict[str, str], filepath: str) -> bool:
        """
        Экспорт данных в PDF
        
        Args:
            db: объект Database
            tables_to_export: dict {display_name: table_name}
            filepath: путь к файлу
        """
        try:
            # Регистрируем шрифт
            font_name = ExportImportHelper.register_fonts()
            
            # Создаем PDF
            doc = SimpleDocTemplate(
                filepath,
                pagesize=landscape(A4),
                rightMargin=20,
                leftMargin=20,
                topMargin=20,
                bottomMargin=20
            )
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Кастомные стили
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=font_name,
                fontSize=18,
                textColor=colors.HexColor('#1A529C'),
                spaceAfter=12,
                alignment=1  # Center
            )
            
            table_header_style = ParagraphStyle(
                'TableHeader',
                parent=styles['Normal'],
                fontName=f'{font_name}-Bold',
                fontSize=9,
                textColor=colors.white,
                alignment=1  # Center
            )
            
            table_cell_style = ParagraphStyle(
                'TableCell',
                parent=styles['Normal'],
                fontName=font_name,
                fontSize=8,
                textColor=colors.black,
                alignment=0  # Left
            )
            
            for display_name, table_name in tables_to_export.items():
                # Получаем данные
                rows = db.get_table_data(table_name)
                if not rows:
                    continue
                
                # Заголовок таблицы
                elements.append(Paragraph(f"{display_name}", title_style))
                elements.append(Spacer(1, 8))
                
                # Подготовка данных для таблицы
                headers = list(rows[0].keys())
                table_data = [headers]
                
                for row in rows:
                    row_values = []
                    for header in headers:
                        value = ExportImportHelper.convert_value_for_excel(row.get(header))
                        # Ограничиваем длину текста
                        if len(str(value)) > 60:
                            value = str(value)[:57] + "..."
                        row_values.append(Paragraph(str(value), table_cell_style))
                    table_data.append(row_values)
                
                # Создаем таблицу
                col_widths = [inch * 1.5] * len(headers)
                table = Table(table_data, colWidths=col_widths)
                
                # Стилизация таблицы
                table.setStyle(TableStyle([
                    # Заголовок
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A529C')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), f'{font_name}-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    
                    # Данные
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 1), (-1, -1), font_name),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('TOPPADDING', (0, 1), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                    
                    # Сетка
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    
                    # Чередование строк
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FAFAFA')),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FAFAFA')]),
                ]))
                
                elements.append(table)
                elements.append(Spacer(1, 20))
                elements.append(PageBreak())
            
            # Сохраняем
            doc.build(elements)
            return True
            
        except Exception as e:
            print(f"❌ Ошибка экспорта в PDF: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    @staticmethod
    def import_from_xlsx(db, table_name: str, filepath: str, skip_duplicates: bool = True) -> Dict[str, Any]:
        """
        Импорт данных из Excel
        
        Args:
            db: объект Database
            table_name: имя таблицы для импорта
            filepath: путь к файлу
            skip_duplicates: пропускать ли дубликаты
            
        Returns:
            dict со статистикой импорта
        """
        stats = {
            'total': 0,
            'imported': 0,
            'skipped': 0,
            'errors': 0,
            'error_messages': []
        }
        
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Получаем заголовки
            headers = [cell.value for cell in ws[1]]
            
            # Получаем существующие записи для проверки дубликатов
            existing_records = set()
            if skip_duplicates and table_name in ['materials', 'users', 'categories']:
                # Определяем уникальное поле
                unique_field = {
                    'materials': 'name',
                    'users': 'login',
                    'categories': 'name'
                }.get(table_name, 'id')
                
                conn = db.get_connection()
                cursor = conn.cursor()
                cursor.execute(f"SELECT {unique_field} FROM {table_name}")
                existing_records = {str(row[0]) for row in cursor.fetchall()}
                cursor.close()
                conn.close()
            
            # Обрабатываем строки
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                stats['total'] += 1
                
                try:
                    # Пропускаем пустые строки
                    if all(cell is None for cell in row):
                        stats['skipped'] += 1
                        continue
                    
                    # Создаем словарь данных
                    record_data = {}
                    for col_idx, value in enumerate(row):
                        if col_idx < len(headers) and headers[col_idx] and value is not None:
                            header = headers[col_idx]
                            
                            # Конвертируем значения
                            if isinstance(value, (datetime, date)):
                                record_data[header] = value
                            elif isinstance(value, bool):
                                record_data[header] = value
                            else:
                                # Пытаемся определить тип
                                str_value = str(value).strip()
                                if str_value.lower() in ('none', 'null', ''):
                                    record_data[header] = None
                                else:
                                    record_data[header] = str_value
                    
                    # Проверяем на дубликаты
                    if skip_duplicates:
                        unique_field = {
                            'materials': 'name',
                            'users': 'login',
                            'categories': 'name'
                        }.get(table_name)
                        
                        if unique_field and unique_field in record_data:
                            if str(record_data[unique_field]) in existing_records:
                                stats['skipped'] += 1
                                continue
                    
                    # Вставляем запись
                    if record_data:  # Только если есть данные
                        db.insert_record(table_name, record_data)
                        stats['imported'] += 1
                        
                        # Добавляем в существующие для следующих проверок
                        if skip_duplicates:
                            unique_field = {
                                'materials': 'name',
                                'users': 'login',
                                'categories': 'name'
                            }.get(table_name)
                            if unique_field and unique_field in record_data:
                                existing_records.add(str(record_data[unique_field]))
                
                except Exception as e:
                    stats['errors'] += 1
                    stats['error_messages'].append(f"Строка {row_idx}: {str(e)}")
            
            return stats
            
        except Exception as e:
            stats['errors'] += 1
            stats['error_messages'].append(f"Ошибка файла: {str(e)}")
            return stats